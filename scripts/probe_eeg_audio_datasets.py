#!/usr/bin/env python3
"""Lightweight probes for EEG-audio public datasets.

The script intentionally avoids downloading full EEG archives. It validates
dataset availability by reading BIDS metadata, annotations, small metadata
files, WAV headers, and the first bytes of representative raw EEG files.
"""

from __future__ import annotations

import argparse
import csv
import io
import hashlib
import json
import re
import struct
import sys
import time
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import requests

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


S3_BASE = "https://s3.amazonaws.com/openneuro.org"
S3_NS = {"s": "http://s3.amazonaws.com/doc/2006-03-01/"}


@dataclass
class Target:
    label: str
    url: str
    kind: str
    max_bytes: int | None = None


@dataclass
class Probe:
    dataset_id: str
    title: str
    source: str
    url: str
    priority: str
    fit: str
    targets: list[Target] = field(default_factory=list)
    zenodo_id: int | None = None
    github_api: str | None = None


def s3_url(key: str) -> str:
    return f"{S3_BASE}/{key}"


def get_bytes(
    url: str,
    max_bytes: int | None = None,
    timeout: int = 30,
    attempts: int = 4,
) -> tuple[bytes, dict[str, str]]:
    headers = {}
    if max_bytes is not None:
        headers["Range"] = f"bytes=0-{max_bytes - 1}"
    last_exc: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=timeout)
            resp.raise_for_status()
            return resp.content, dict(resp.headers)
        except (requests.RequestException, requests.Timeout) as exc:
            last_exc = exc
            if attempt == attempts:
                break
            time.sleep(0.75 * attempt)
    assert last_exc is not None
    raise last_exc


def list_s3(prefix: str, max_keys: int = 1000) -> dict[str, Any]:
    resp = requests.get(
        S3_BASE,
        params={"list-type": "2", "prefix": prefix, "max-keys": str(max_keys)},
        timeout=30,
    )
    resp.raise_for_status()
    root = ET.fromstring(resp.text)
    files = []
    for elem in root.findall("s:Contents", S3_NS):
        files.append(
            {
                "key": elem.findtext("s:Key", namespaces=S3_NS),
                "size": int(elem.findtext("s:Size", namespaces=S3_NS) or 0),
            }
        )
    return {
        "prefix": prefix,
        "is_truncated": root.findtext("s:IsTruncated", namespaces=S3_NS) == "true",
        "key_count_page": int(root.findtext("s:KeyCount", namespaces=S3_NS) or 0),
        "files": files,
    }


def parse_wav_header(blob: bytes) -> dict[str, Any]:
    if len(blob) < 44 or blob[0:4] != b"RIFF" or blob[8:12] != b"WAVE":
        return {"valid_wav": False, "magic": blob[:12].hex()}
    pos = 12
    fmt = None
    data_size = None
    while pos + 8 <= len(blob):
        chunk_id = blob[pos : pos + 4]
        chunk_size = struct.unpack("<I", blob[pos + 4 : pos + 8])[0]
        chunk_start = pos + 8
        if chunk_id == b"fmt " and chunk_start + min(chunk_size, 16) <= len(blob):
            fmt_data = blob[chunk_start : chunk_start + min(chunk_size, 16)]
            if len(fmt_data) >= 16:
                audio_format, channels, sample_rate, byte_rate, block_align, bits = struct.unpack(
                    "<HHIIHH", fmt_data[:16]
                )
                fmt = {
                    "audio_format": audio_format,
                    "channels": channels,
                    "sample_rate": sample_rate,
                    "byte_rate": byte_rate,
                    "block_align": block_align,
                    "bits_per_sample": bits,
                }
        if chunk_id == b"data":
            data_size = chunk_size
            break
        pos = chunk_start + chunk_size + (chunk_size % 2)
    out = {"valid_wav": True}
    if fmt:
        out.update(fmt)
    if data_size is not None and fmt and fmt.get("byte_rate"):
        out["duration_sec_est"] = round(data_size / fmt["byte_rate"], 3)
    return out


def preview_table(blob: bytes, delimiter: str | None = None, max_rows: int = 5) -> dict[str, Any]:
    text = blob.decode("utf-8", errors="replace")
    sample = text[:4096]
    if delimiter is None:
        delimiter = "\t" if "\t" in sample.splitlines()[0] else ","
    rows = list(csv.reader(io.StringIO(sample), delimiter=delimiter))
    return {
        "delimiter": "\\t" if delimiter == "\t" else delimiter,
        "columns": rows[0] if rows else [],
        "preview_rows": rows[1 : 1 + max_rows] if len(rows) > 1 else [],
        "line_count_in_sample": len(sample.splitlines()),
    }


def preview_textgrid(blob: bytes) -> dict[str, Any]:
    text = blob.decode("utf-8", errors="replace")
    labels = re.findall(r'text\s*=\s*"([^"]*)"', text)
    nonempty = [label for label in labels if label.strip()]
    xmax = re.findall(r"xmax\s*=\s*([0-9.]+)", text)
    return {
        "interval_labels": len(labels),
        "nonempty_labels": len(nonempty),
        "first_nonempty": nonempty[:10],
        "duration_sec_hint": float(xmax[0]) if xmax else None,
    }


def preview_xlsx(blob: bytes) -> dict[str, Any]:
    if openpyxl is None:
        return {"error": "openpyxl unavailable"}
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    out = {"sheets": wb.sheetnames}
    first = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(first.iter_rows(values_only=True), start=1):
        rows.append([str(x) if x is not None else "" for x in row[:8]])
        if i >= 5:
            break
    out["first_sheet"] = first.title
    out["max_row"] = first.max_row
    out["max_column"] = first.max_column
    out["preview_rows"] = rows
    return out


def preview_zip(blob: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        infos = zf.infolist()
        return {
            "entries": len(infos),
            "first_entries": [
                {"filename": info.filename, "size": info.file_size} for info in infos[:12]
            ],
        }


def preview_binary(blob: bytes, headers: dict[str, str]) -> dict[str, Any]:
    ascii_prefix = "".join(chr(b) if 32 <= b < 127 else "." for b in blob[:64])
    return {
        "bytes_read": len(blob),
        "content_range": headers.get("Content-Range"),
        "magic_hex": blob[:16].hex(),
        "ascii_prefix": ascii_prefix,
    }


def artifact_extension(kind: str) -> str:
    return {
        "json": ".json",
        "tsv": ".tsv",
        "csv": ".csv",
        "text": ".txt",
        "textgrid": ".TextGrid",
        "wav": ".wav.header.bin",
        "xlsx": ".xlsx",
        "zip": ".zip",
        "binary": ".bin",
    }.get(kind, ".bin")


def safe_filename(text: str) -> str:
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", text.strip()).strip("._")
    return name or "artifact"


def save_artifact(
    artifact_dir: Path | None,
    dataset_id: str,
    target_index: int,
    target: Target,
    blob: bytes,
    headers: dict[str, str],
) -> dict[str, Any]:
    sha = hashlib.sha256(blob).hexdigest()
    info = {
        "sha256": sha,
        "is_partial": target.max_bytes is not None or headers.get("Content-Range") is not None,
    }
    if artifact_dir is None:
        return info

    dataset_dir = artifact_dir / safe_filename(dataset_id)
    dataset_dir.mkdir(parents=True, exist_ok=True)
    path = dataset_dir / f"{target_index:02d}_{safe_filename(target.label)}{artifact_extension(target.kind)}"
    path.write_bytes(blob)
    info["artifact_path"] = str(path)
    info["artifact_bytes"] = len(blob)
    return info


def summarize_target(
    target: Target,
    dataset_id: str,
    target_index: int,
    artifact_dir: Path | None = None,
) -> dict[str, Any]:
    blob, headers = get_bytes(target.url, target.max_bytes)
    artifact = save_artifact(artifact_dir, dataset_id, target_index, target, blob, headers)
    out: dict[str, Any] = {
        "label": target.label,
        "kind": target.kind,
        "url": target.url,
        "bytes_read": len(blob),
        "content_length_header": headers.get("Content-Length"),
        "content_range": headers.get("Content-Range"),
        "artifact": artifact,
    }
    try:
        if target.kind == "json":
            out["parsed"] = json.loads(blob.decode("utf-8"))
        elif target.kind == "tsv":
            out["parsed"] = preview_table(blob, "\t")
        elif target.kind == "csv":
            out["parsed"] = preview_table(blob, ",")
        elif target.kind == "text":
            text = blob.decode("utf-8", errors="replace")
            out["parsed"] = {"lines": text.splitlines()[:12], "char_count": len(text)}
        elif target.kind == "textgrid":
            out["parsed"] = preview_textgrid(blob)
        elif target.kind == "wav":
            out["parsed"] = parse_wav_header(blob)
        elif target.kind == "xlsx":
            out["parsed"] = preview_xlsx(blob)
        elif target.kind == "zip":
            out["parsed"] = preview_zip(blob)
        else:
            out["parsed"] = preview_binary(blob, headers)
    except Exception as exc:
        out["parsed"] = {"error": f"{type(exc).__name__}: {exc}"}
    return out


def safe_summarize_target(
    target: Target,
    dataset_id: str,
    target_index: int,
    artifact_dir: Path | None = None,
) -> dict[str, Any]:
    try:
        return summarize_target(target, dataset_id, target_index, artifact_dir)
    except Exception as exc:
        return {
            "label": target.label,
            "kind": target.kind,
            "url": target.url,
            "error": f"{type(exc).__name__}: {exc}",
        }


def summarize_targets(probe: Probe, artifact_dir: Path | None = None) -> list[dict[str, Any]]:
    return [
        safe_summarize_target(target, probe.dataset_id, index, artifact_dir)
        for index, target in enumerate(probe.targets, start=1)
    ]


def probe_openneuro(probe: Probe, artifact_dir: Path | None = None) -> dict[str, Any]:
    ds = probe.dataset_id
    summary = {
        "dataset_id": ds,
        "title": probe.title,
        "source": probe.source,
        "url": probe.url,
        "priority": probe.priority,
        "fit": probe.fit,
    }
    page = list_s3(f"{ds}/", 1000)
    summary["s3_first_page"] = {
        "key_count_page": page["key_count_page"],
        "is_truncated": page["is_truncated"],
        "sample_small_files": [
            f for f in page["files"][:25] if f["size"] < 1_000_000
        ][:10],
    }
    summary["targets"] = summarize_targets(probe, artifact_dir)
    return summary


def probe_zenodo(probe: Probe, artifact_dir: Path | None = None) -> dict[str, Any]:
    assert probe.zenodo_id is not None
    resp = requests.get(f"https://zenodo.org/api/records/{probe.zenodo_id}", timeout=30)
    resp.raise_for_status()
    record = resp.json()
    files = record.get("files", [])
    summary = {
        "dataset_id": probe.dataset_id,
        "title": record.get("metadata", {}).get("title", probe.title),
        "source": probe.source,
        "url": probe.url,
        "priority": probe.priority,
        "fit": probe.fit,
        "doi": record.get("doi"),
        "file_count": len(files),
        "files_preview": [
            {
                "key": f.get("key"),
                "size": f.get("size"),
                "url": f.get("links", {}).get("self"),
            }
            for f in files[:12]
        ],
    }
    summary["targets"] = summarize_targets(probe, artifact_dir)
    return summary


def probe_github(probe: Probe, artifact_dir: Path | None = None) -> dict[str, Any]:
    assert probe.github_api is not None
    resp = requests.get(probe.github_api, timeout=30)
    resp.raise_for_status()
    listing = resp.json()
    summary = {
        "dataset_id": probe.dataset_id,
        "title": probe.title,
        "source": probe.source,
        "url": probe.url,
        "priority": probe.priority,
        "fit": probe.fit,
        "github_listing": [
            {
                "name": item.get("name"),
                "type": item.get("type"),
                "size": item.get("size"),
                "download_url": item.get("download_url"),
            }
            for item in listing
        ],
    }
    summary["targets"] = summarize_targets(probe, artifact_dir)
    return summary


def build_probes() -> list[Probe]:
    return [
        Probe(
            "ds004408",
            "EEG responses to continuous naturalistic speech",
            "OpenNeuro",
            "https://openneuro.org/datasets/ds004408",
            "core",
            "Natural speech pretraining; word/phoneme TextGrid alignment",
            [
                Target("dataset_description", s3_url("ds004408/dataset_description.json"), "json"),
                Target("participants", s3_url("ds004408/participants.tsv"), "tsv"),
                Target("audio01 TextGrid", s3_url("ds004408/stimuli/audio01.TextGrid"), "textgrid"),
                Target("audio01 wav header", s3_url("ds004408/stimuli/audio01.wav"), "wav", 4096),
                Target("run01 EEG sidecar", s3_url("ds004408/sub-001/eeg/sub-001_task-listening_run-01_eeg.json"), "json"),
                Target("run01 channels", s3_url("ds004408/sub-001/eeg/sub-001_task-listening_run-01_channels.tsv"), "tsv"),
                Target("run01 BrainVision header", s3_url("ds004408/sub-001/eeg/sub-001_task-listening_run-01_eeg.vhdr"), "text"),
                Target("run01 raw EEG bytes", s3_url("ds004408/sub-001/eeg/sub-001_task-listening_run-01_eeg.eeg"), "binary", 512),
            ],
        ),
        Probe(
            "ds005345",
            "Le Petit Prince Multi-talker",
            "OpenNeuro",
            "https://openneuro.org/datasets/ds005345",
            "core",
            "Natural Mandarin speech; single/multi-talker semantic and acoustic alignment",
            [
                Target("dataset_description", s3_url("ds005345/dataset_description.json"), "json"),
                Target("female word info", s3_url("ds005345/annotation/single_female_word_information.csv"), "csv"),
                Target("female acoustic sample", s3_url("ds005345/annotation/single_female_acoustic.csv"), "csv", 262144),
                Target("single female wav header", s3_url("ds005345/stimuli/single_female.wav"), "wav", 4096),
                Target("raw EEG sidecar", s3_url("ds005345/sub-01/eeg/sub-01_task-multitalker_eeg.json"), "json"),
                Target("raw EEG BrainVision header", s3_url("ds005345/sub-01/eeg/sub-01_task-multitalker_eeg.vhdr"), "text"),
                Target("raw EEG bytes", s3_url("ds005345/sub-01/eeg/sub-01_task-multitalker_eeg.eeg"), "binary", 512),
                Target("preprocessed fif bytes", s3_url("ds005345/derivatives/sub-01/eeg/sub-01_task-multitalker_run-1_eeg_preprocessed.fif"), "binary", 512),
            ],
        ),
        Probe(
            "ds004718",
            "LPPHK older Cantonese speakers",
            "OpenNeuro",
            "https://openneuro.org/datasets/ds004718",
            "core",
            "Best word/prosody alignment; older Cantonese natural speech",
            [
                Target("dataset_description", s3_url("ds004718/dataset_description.json"), "json"),
                Target("sentences", s3_url("ds004718/sourcedata/annotation/snts.txt"), "text"),
                Target("word timing xlsx", s3_url("ds004718/sourcedata/annotation/lppHK_timing_word_information.xlsx"), "xlsx"),
                Target("trigger sentence xlsx", s3_url("ds004718/sourcedata/annotation/EEG_trigger_and_sentence_number.xlsx"), "xlsx"),
                Target("acoustic csv sample", s3_url("ds004718/sourcedata/annotation/wav_acoustic.csv"), "csv", 262144),
                Target("sentence wav header", s3_url("ds004718/sourcedata/stimuli/audio_files_segmented_by_sentence/Part%201/1.003.wav"), "wav", 4096),
                Target("raw EEG sidecar", s3_url("ds004718/sub-HK001/eeg/sub-HK001_task-lppHK_eeg.json"), "json"),
                Target("raw EEG EEGLAB set bytes", s3_url("ds004718/sub-HK001/eeg/sub-HK001_task-lppHK_eeg.set"), "binary", 512),
            ],
        ),
        Probe(
            "ds006104",
            "EEG dataset for speech decoding",
            "OpenNeuro",
            "https://openneuro.org/datasets/ds006104",
            "controlled",
            "Phoneme/articulation probe; not continuous natural speech",
            [
                Target("dataset_description", s3_url("ds006104/dataset_description.json"), "json"),
                Target("channels", s3_url("ds006104/sub-P01/ses-01/eeg/sub-P01_ses-01_task-phonemes_channels.tsv"), "tsv"),
                Target("events", s3_url("ds006104/sub-P01/ses-01/eeg/sub-P01_ses-01_task-phonemes_events.tsv"), "tsv"),
                Target("EEG sidecar", s3_url("ds006104/sub-P01/ses-01/eeg/sub-P01_ses-01_task-phonemes_eeg.json"), "json"),
                Target("raw EDF bytes", s3_url("ds006104/sub-P01/ses-01/eeg/sub-P01_ses-01_task-phonemes_eeg.edf"), "binary", 512),
            ],
        ),
        Probe(
            "ds006434",
            "ABR to natural speech and selective attention",
            "OpenNeuro",
            "https://openneuro.org/datasets/ds006434",
            "alignment",
            "High-rate timing/stimulus-alignment stress test",
            [
                Target("dataset_description", s3_url("ds006434/dataset_description.json"), "json"),
                Target("events", s3_url("ds006434/sub-dichotic02/eeg/sub-dichotic02_task-exp2DichoticCortex_events.tsv"), "tsv"),
                Target("channels", s3_url("ds006434/sub-dichotic02/eeg/sub-dichotic02_task-exp2DichoticCortex_channels.tsv"), "tsv"),
                Target("EEG sidecar", s3_url("ds006434/sub-dichotic02/eeg/sub-dichotic02_task-exp2DichoticCortex_eeg.json"), "json"),
                Target("stim wav header", s3_url("ds006434/stimuli/exp2Dichotic/wrinkle_alchemyst000.wav"), "wav", 4096),
                Target("raw EEG bytes", s3_url("ds006434/sub-dichotic02/eeg/sub-dichotic02_task-exp2DichoticCortex_eeg.eeg"), "binary", 512),
            ],
        ),
        Probe(
            "ds003774",
            "MUSIN-G music listening EEG",
            "OpenNeuro",
            "https://openneuro.org/datasets/ds003774",
            "music",
            "Natural music self-supervised EEG tokenization; weak affect labels",
            [
                Target("dataset_description", s3_url("ds003774/dataset_description.json"), "json"),
                Target("events", s3_url("ds003774/sourcedata/sub-001/eeg/sub-001_task-ListeningandResponse_events.tsv"), "tsv"),
                Target("channels", s3_url("ds003774/sourcedata/sub-001/eeg/sub-001_task-ListeningandResponse_channels.tsv"), "tsv"),
                Target("EEG sidecar", s3_url("ds003774/sourcedata/sub-001/eeg/sub-001_task-ListeningandResponse_eeg.json"), "json"),
                Target("song wav header", s3_url("ds003774/Code/ESongs/1.esh.wav"), "wav", 4096),
                Target("raw EEG set bytes", s3_url("ds003774/sourcedata/sub-001/eeg/sub-001_task-ListeningandResponse_eeg.set"), "binary", 512),
            ],
        ),
        Probe(
            "ds007591",
            "Delineating neural contributions to EEG-based speech decoding",
            "OpenNeuro",
            "https://openneuro.org/datasets/ds007591",
            "secondary",
            "Speech production/covert speech sanity check; small subject count",
            [
                Target("dataset_description", s3_url("ds007591/dataset_description.json"), "json"),
                Target("participants", s3_url("ds007591/participants.tsv"), "tsv"),
                Target("events", s3_url("ds007591/sub-1/ses-20230511/eeg/sub-1_ses-20230511_task-minimallyovert_acq-calibration_run-01_events.tsv"), "tsv"),
                Target("channels", s3_url("ds007591/sub-1/ses-20230511/eeg/sub-1_ses-20230511_task-minimallyovert_acq-calibration_run-01_channels.tsv"), "tsv"),
                Target("EEG sidecar", s3_url("ds007591/sub-1/ses-20230511/eeg/sub-1_ses-20230511_task-minimallyovert_acq-calibration_run-01_eeg.json"), "json"),
                Target("raw EDF bytes", s3_url("ds007591/sub-1/ses-20230511/eeg/sub-1_ses-20230511_task-minimallyovert_acq-calibration_run-01_eeg.edf"), "binary", 512),
            ],
        ),
        Probe(
            "zenodo-4004271",
            "KULeuven auditory attention detection",
            "Zenodo",
            "https://zenodo.org/records/4004271",
            "baseline",
            "Classic envelope/AAD benchmark; use strict trial/story split",
            [
                Target("README", "https://zenodo.org/api/records/4004271/files/README.txt.txt/content", "text"),
                Target("preprocess script", "https://zenodo.org/api/records/4004271/files/preprocess_data.m/content", "text"),
            ],
            zenodo_id=4004271,
        ),
        Probe(
            "zenodo-1199011",
            "DTU EEG and audio dataset for AAD",
            "Zenodo",
            "https://zenodo.org/records/1199011",
            "baseline",
            "Reverberant competing speech; good robustness test",
            [
                Target("preproc script", "https://zenodo.org/api/records/1199011/files/preproc_data.m/content", "text"),
            ],
            zenodo_id=1199011,
        ),
        Probe(
            "zenodo-4518754",
            "Ultra-high-density 255-channel EEG-AAD",
            "Zenodo",
            "https://zenodo.org/records/4518754",
            "spatial",
            "High-density spatial tokenizer and sensor-ablation benchmark",
            [
                Target("misc zip", "https://zenodo.org/api/records/4518754/files/misc.zip/content", "zip"),
                Target("scripts zip", "https://zenodo.org/api/records/4518754/files/scripts.zip/content", "zip"),
            ],
            zenodo_id=4518754,
        ),
        Probe(
            "zenodo-7078451",
            "ESAA Mandarin EEG-Speech AAD",
            "Zenodo",
            "https://zenodo.org/records/7078451",
            "secondary",
            "Mandarin tonal-language AAD; useful after core OpenNeuro sets",
            [
                Target("readme", "https://zenodo.org/api/records/7078451/files/readme.txt/content", "text"),
                Target("preprocess zip", "https://zenodo.org/api/records/7078451/files/preprocess.zip/content", "zip"),
                Target("baseline zip", "https://zenodo.org/api/records/7078451/files/cnn_baseline.zip/content", "zip"),
            ],
            zenodo_id=7078451,
        ),
        Probe(
            "openmiir",
            "OpenMIIR music perception and imagination",
            "GitHub/Figshare",
            "https://github.com/sstober/openmiir",
            "music",
            "Beat/tempo/meter alignment probe before larger natural music sets",
            [
                Target("stimuli metadata", "https://raw.githubusercontent.com/sstober/openmiir/master/meta/Stimuli_Meta.v2.xlsx", "xlsx"),
                Target("electrode info", "https://raw.githubusercontent.com/sstober/openmiir/master/meta/electrode_info.xlsx", "xlsx"),
                Target("beat annotations", "https://raw.githubusercontent.com/sstober/openmiir/master/meta/beats.v2/1_beats.txt", "text"),
            ],
            github_api="https://api.github.com/repos/sstober/openmiir/contents/meta",
        ),
        Probe(
            "zenodo-4537751",
            "MAD-EEG target-instrument music attention",
            "Zenodo",
            "https://zenodo.org/records/4537751",
            "music-secondary",
            "Small music AAD set for target-instrument attention",
            [
                Target("behavioral data", "https://zenodo.org/api/records/4537751/files/behavioural_data.xlsx/content", "xlsx"),
                Target("raw yaml", "https://zenodo.org/api/records/4537751/files/madeeg_raw.yaml/content", "text", 65536),
                Target("sequences yaml", "https://zenodo.org/api/records/4537751/files/madeeg_sequences_raw.yaml/content", "text", 65536),
            ],
            zenodo_id=4537751,
        ),
    ]


def run(probes: list[Probe], artifact_dir: Path | None = None) -> list[dict[str, Any]]:
    results = []
    for probe in probes:
        print(f"Probing {probe.dataset_id}...", file=sys.stderr)
        try:
            if probe.source == "OpenNeuro":
                results.append(probe_openneuro(probe, artifact_dir))
            elif probe.source == "Zenodo":
                results.append(probe_zenodo(probe, artifact_dir))
            else:
                results.append(probe_github(probe, artifact_dir))
        except Exception as exc:
            results.append(
                {
                    "dataset_id": probe.dataset_id,
                    "title": probe.title,
                    "source": probe.source,
                    "url": probe.url,
                    "priority": probe.priority,
                    "fit": probe.fit,
                    "error": f"{type(exc).__name__}: {exc}",
                }
            )
    return results


def compact_markdown(results: list[dict[str, Any]]) -> str:
    lines = [
        "# EEG-audio dataset probe results",
        "",
        "Generated by `scripts/probe_eeg_audio_datasets.py`. Full raw EEG archives were not downloaded; checks use metadata, annotations, WAV headers, and HTTP byte ranges.",
        "",
        "| Dataset | Source | Priority | Probe status | Evidence |",
        "| --- | --- | --- | --- | --- |",
    ]
    for res in results:
        if "error" in res:
            status = "ERROR"
            evidence = res["error"]
        else:
            status = "OK"
            bits = []
            for target in res.get("targets", [])[:4]:
                parsed = target.get("parsed", {})
                if isinstance(parsed, dict):
                    if target["kind"] == "wav":
                        bits.append(
                            f"{target['label']}: {parsed.get('sample_rate')} Hz, {parsed.get('duration_sec_est')} s"
                        )
                    elif target["kind"] in {"csv", "tsv"}:
                        bits.append(
                            f"{target['label']}: {len(parsed.get('columns', []))} columns"
                        )
                    elif target["kind"] == "textgrid":
                        bits.append(
                            f"{target['label']}: {parsed.get('nonempty_labels')} labels"
                        )
                    elif target["kind"] == "xlsx":
                        bits.append(
                            f"{target['label']}: sheets={','.join(parsed.get('sheets', [])[:3])}"
                        )
                    elif target["kind"] == "zip":
                        bits.append(
                            f"{target['label']}: {parsed.get('entries')} zip entries"
                        )
                    elif target["kind"] == "json":
                        value = parsed.get("Name") or parsed.get("SamplingFrequency") or "json"
                        bits.append(f"{target['label']}: {markdown_cell(str(value))}")
                    else:
                        bits.append(f"{target['label']}: {target.get('bytes_read')} bytes")
            evidence = "; ".join(bits)
        lines.append(
            f"| {res.get('dataset_id')} | {res.get('source')} | {res.get('priority')} | {status} | {evidence} |"
        )
    lines.append("")
    return "\n".join(lines)


def markdown_cell(value: Any) -> str:
    text = str(value).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip().replace("|", "\\|")


def summarize_parsed_value(target: dict[str, Any]) -> str:
    parsed = target.get("parsed", {})
    if "error" in target:
        return target["error"]
    if not isinstance(parsed, dict):
        return "unparsed"
    kind = target.get("kind")
    if kind == "json":
        keys = ["Name", "DatasetDOI", "License", "SamplingFrequency", "EEGReference", "Manufacturer"]
        values = [f"{key}={markdown_cell(parsed[key])}" for key in keys if key in parsed]
        return "; ".join(values) or "json parsed"
    if kind in {"csv", "tsv"}:
        return f"columns={parsed.get('columns', [])}; preview_rows={parsed.get('preview_rows', [])[:2]}"
    if kind == "xlsx":
        return (
            f"sheets={parsed.get('sheets', [])}; "
            f"shape=({parsed.get('max_row')}, {parsed.get('max_column')}); "
            f"preview_rows={parsed.get('preview_rows', [])[:2]}"
        )
    if kind == "textgrid":
        return (
            f"nonempty_labels={parsed.get('nonempty_labels')}; "
            f"duration_sec_hint={parsed.get('duration_sec_hint')}; "
            f"first_nonempty={parsed.get('first_nonempty', [])[:10]}"
        )
    if kind == "wav":
        return (
            f"sample_rate={parsed.get('sample_rate')}; channels={parsed.get('channels')}; "
            f"bits={parsed.get('bits_per_sample')}; duration_sec_est={parsed.get('duration_sec_est')}"
        )
    if kind == "zip":
        return f"entries={parsed.get('entries')}; first_entries={parsed.get('first_entries', [])[:5]}"
    return f"bytes={target.get('bytes_read')}; magic={parsed.get('magic_hex')}; range={parsed.get('content_range')}"


def detailed_markdown(results: list[dict[str, Any]]) -> str:
    lines = [
        "# EEG-audio dataset probe detailed report",
        "",
        "This report is generated from real remote metadata probes. Full EEG archives are not downloaded by default; partial byte-range artifacts are explicitly marked.",
        "",
        "## Dataset Summary",
        "",
        "| Dataset | Source | Priority | Fit | URL |",
        "| --- | --- | --- | --- | --- |",
    ]
    for res in results:
        lines.append(
            f"| {markdown_cell(res.get('dataset_id'))} | {markdown_cell(res.get('source'))} | {markdown_cell(res.get('priority'))} | {markdown_cell(res.get('fit'))} | {markdown_cell(res.get('url'))} |"
        )

    lines.extend(["", "## Target-Level Evidence", ""])
    for res in results:
        lines.extend([f"### {res.get('dataset_id')} - {res.get('title')}", ""])
        if "error" in res:
            lines.extend([f"Probe error: `{res['error']}`", ""])
            continue
        if res.get("doi"):
            lines.append(f"DOI: `{res['doi']}`")
        if res.get("file_count") is not None:
            lines.append(f"Zenodo file count: `{res['file_count']}`")
        if res.get("s3_first_page"):
            page = res["s3_first_page"]
            lines.append(
                f"OpenNeuro S3 first page: `{page.get('key_count_page')}` keys; truncated=`{page.get('is_truncated')}`"
            )
        lines.append("")
        lines.append("| Target | Kind | Bytes | Partial | Artifact | Parsed evidence |")
        lines.append("| --- | --- | ---: | --- | --- | --- |")
        for target in res.get("targets", []):
            artifact = target.get("artifact", {})
            artifact_path = artifact.get("artifact_path", "")
            partial = artifact.get("is_partial", "")
            evidence = markdown_cell(summarize_parsed_value(target))
            lines.append(
                f"| {markdown_cell(target.get('label'))} | {markdown_cell(target.get('kind'))} | {markdown_cell(target.get('bytes_read', ''))} | {markdown_cell(partial)} | {markdown_cell(artifact_path)} | {evidence} |"
            )
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json-out", type=Path)
    parser.add_argument("--md-out", type=Path)
    parser.add_argument("--detail-md-out", type=Path)
    parser.add_argument(
        "--artifact-dir",
        type=Path,
        help="Directory for saved raw probe artifacts: small metadata files, headers, and byte-range samples.",
    )
    parser.add_argument(
        "--only",
        nargs="*",
        help="Optional dataset ids to probe, e.g. ds004408 ds004718.",
    )
    args = parser.parse_args()

    probes = build_probes()
    if args.only:
        wanted = set(args.only)
        probes = [p for p in probes if p.dataset_id in wanted]

    results = run(probes, args.artifact_dir)
    if args.json_out:
        args.json_out.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    markdown = compact_markdown(results)
    if args.md_out:
        args.md_out.write_text(markdown, encoding="utf-8")
    if args.detail_md_out:
        args.detail_md_out.write_text(detailed_markdown(results), encoding="utf-8")
    print(markdown)


if __name__ == "__main__":
    main()
